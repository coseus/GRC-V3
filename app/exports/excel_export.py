from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def generate_excel_report(
    *,
    company_name: str,
    assessment_name: str,
    framework_name: str,
    responses,
    domain_scores,
    executive_summary=None,
    recommendations=None,
):
    wb = Workbook()

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)

    # Sheet principal: Questions & Answers
    ws_answers = wb.active
    ws_answers.title = "Questions_Answers"

    ws_answers["A1"] = "Company"
    ws_answers["B1"] = company_name
    ws_answers["A2"] = "Assessment"
    ws_answers["B2"] = assessment_name
    ws_answers["A3"] = "Framework"
    ws_answers["B3"] = framework_name

    ws_answers["A1"].font = bold_font
    ws_answers["A2"].font = bold_font
    ws_answers["A3"].font = bold_font

    headers = [
        "Domain",
        "Question ID",
        "Question",
        "Answer",
        "Score",
        "Weight",
        "Risk",
        "Notes",
        "Proof",
    ]

    start_row = 5
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_answers.cell(row=start_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, row in enumerate(responses or [], start=start_row + 1):
        ws_answers.cell(row=row_idx, column=1, value=row.get("domain_name") or row.get("domain"))
        ws_answers.cell(row=row_idx, column=2, value=row.get("question_id"))
        ws_answers.cell(row=row_idx, column=3, value=row.get("question_text") or row.get("question"))
        ws_answers.cell(row=row_idx, column=4, value=row.get("answer_value") or row.get("selected_value"))
        ws_answers.cell(row=row_idx, column=5, value=row.get("score"))
        ws_answers.cell(row=row_idx, column=6, value=row.get("weight"))
        ws_answers.cell(row=row_idx, column=7, value=row.get("risk"))
        ws_answers.cell(row=row_idx, column=8, value=row.get("notes") or row.get("comment"))
        ws_answers.cell(row=row_idx, column=9, value=row.get("proof") or row.get("evidence"))

    ws_answers.column_dimensions["A"].width = 28
    ws_answers.column_dimensions["B"].width = 16
    ws_answers.column_dimensions["C"].width = 70
    ws_answers.column_dimensions["D"].width = 18
    ws_answers.column_dimensions["E"].width = 10
    ws_answers.column_dimensions["F"].width = 10
    ws_answers.column_dimensions["G"].width = 12
    ws_answers.column_dimensions["H"].width = 40
    ws_answers.column_dimensions["I"].width = 30

    # Sheet secundar: Domain Scores
    ws_scores = wb.create_sheet("Domain_Scores")
    ws_scores.append(["Domain", "Score"])
    for cell in ws_scores[1]:
        cell.fill = header_fill
        cell.font = header_font

    for domain_name, score in (domain_scores or {}).items():
        ws_scores.append([domain_name, score])

    ws_scores.column_dimensions["A"].width = 40
    ws_scores.column_dimensions["B"].width = 12

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()