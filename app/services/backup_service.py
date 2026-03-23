from __future__ import annotations

import json
import shutil
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from app.db.models import Answer, Recommendation
from app.repositories.answers import AnswerRepository
from app.repositories.executive import ExecutiveRepository
from app.repositories.recommendations import RecommendationRepository
from app.services.auth import require_role


VALID_ANSWER_VALUES = {"Fail", "Partial", "Pass", "NotApplicable", ""}


class BackupService:
    def __init__(self, session):
        self.session = session
        self.answer_repo = AnswerRepository(session)
        self.executive_repo = ExecutiveRepository(session)
        self.recommendation_repo = RecommendationRepository(session)

    def export_assessment_json(self, actor, company, assessment) -> bytes:
        require_role(actor, "viewer")

        answers = self.answer_repo.list_for_assessment(assessment.id)
        executive = self.executive_repo.get_by_assessment_id(assessment.id)
        recommendations = self.recommendation_repo.list_for_assessment(assessment.id)

        payload = {
            "company": {
                "name": getattr(company, "name", None),
                "industry": getattr(company, "industry", None),
                "country": getattr(company, "country", None),
                "size": getattr(company, "size", None),
            },
            "assessment": {
                "name": getattr(assessment, "name", None),
                "framework_code": getattr(assessment, "framework_code", None),
                "framework_name": getattr(assessment, "framework_name", None),
                "framework_version": getattr(assessment, "framework_version", "1.0"),
                "status": getattr(assessment, "status", "draft"),
            },
            "answers": [
                {
                    "question_code": a.question_code,
                    "question_text": a.question_text,
                    "domain_code": a.domain_code,
                    "domain_name": a.domain_name,
                    "selected_value": a.selected_value,
                    "score": a.score,
                    "max_score": a.max_score,
                    "weight": a.weight,
                    "comment": a.comment,
                    "evidence": a.evidence,
                    "status": a.status,
                }
                for a in answers
            ],
            "executive_summary": {
                "summary_text": getattr(executive, "summary_text", None) if executive else None,
                "strengths_text": getattr(executive, "strengths_text", None) if executive else None,
                "gaps_text": getattr(executive, "gaps_text", None) if executive else None,
                "recommendations_text": getattr(executive, "recommendations_text", None) if executive else None,
            },
            "recommendations": [
                {
                    "domain_code": r.domain_code,
                    "domain_name": r.domain_name,
                    "question_code": r.question_code,
                    "title": r.title,
                    "description": r.description,
                    "priority": r.priority,
                    "status": r.status,
                    "source": r.source,
                    "score": r.score,
                }
                for r in recommendations
            ],
        }

        return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")

    def import_assessment_json(self, actor, *, company, assessment, json_bytes: bytes):
        require_role(actor, "auditor")

        payload = json.loads(json_bytes.decode("utf-8"))

        answers = payload.get("answers", [])
        executive_payload = payload.get("executive_summary", {})
        recommendations_payload = payload.get("recommendations", [])

        self.session.query(Answer).filter(
            Answer.assessment_id == assessment.id
        ).delete(synchronize_session=False)

        self.session.query(Recommendation).filter(
            Recommendation.assessment_id == assessment.id
        ).delete(synchronize_session=False)

        for item in answers:
            self.answer_repo.upsert(
                assessment_id=assessment.id,
                question_code=item.get("question_code"),
                question_text=item.get("question_text"),
                domain_code=item.get("domain_code"),
                domain_name=item.get("domain_name"),
                selected_value=item.get("selected_value"),
                score=item.get("score"),
                max_score=item.get("max_score"),
                weight=item.get("weight", 1.0),
                comment=item.get("comment"),
                evidence=item.get("evidence"),
                answered_by=getattr(actor, "id", None),
                answered_at=None,
            )

        executive = self.executive_repo.upsert(
            assessment_id=assessment.id,
            summary_text=executive_payload.get("summary_text"),
            strengths_text=executive_payload.get("strengths_text"),
            gaps_text=executive_payload.get("gaps_text"),
            recommendations_text=executive_payload.get("recommendations_text"),
            updated_by=getattr(actor, "id", None),
        )

        for item in recommendations_payload:
            self.recommendation_repo.create(
                assessment_id=assessment.id,
                domain_code=item.get("domain_code"),
                domain_name=item.get("domain_name"),
                question_code=item.get("question_code"),
                title=item.get("title"),
                description=item.get("description"),
                priority=item.get("priority", "medium"),
                status=item.get("status", "open"),
                source=item.get("source", "import"),
                score=item.get("score"),
                updated_by=getattr(actor, "id", None),
            )

        self.session.commit()
        return executive

    def _read_excel_rows(self, excel_bytes: bytes):
        workbook = load_workbook(filename=BytesIO(excel_bytes), data_only=True)

        if "Questions_Answers" not in workbook.sheetnames:
            raise ValueError("Fisierul Excel trebuie sa contina sheet-ul 'Questions_Answers'.")

        ws = workbook["Questions_Answers"]

        header_row_idx = None
        headers = {}

        for row_idx in range(1, min(ws.max_row, 20) + 1):
            values = [ws.cell(row=row_idx, column=col).value for col in range(1, 20)]
            normalized = [str(v).strip() if v is not None else "" for v in values]

            if "Question ID" in normalized and "Answer" in normalized:
                header_row_idx = row_idx
                headers = {str(v).strip(): i + 1 for i, v in enumerate(values) if v is not None}
                break

        if not header_row_idx:
            raise ValueError("Nu am gasit antetul pentru sheet-ul 'Questions_Answers'.")

        required_headers = {"Domain", "Question ID", "Question", "Answer", "Score", "Weight", "Notes", "Proof"}
        missing = [h for h in required_headers if h not in headers]
        if missing:
            raise ValueError(f"Lipsesc coloane obligatorii in Excel: {', '.join(missing)}")

        def get_cell(row_num: int, column_name: str):
            col_idx = headers.get(column_name)
            if not col_idx:
                return None
            return ws.cell(row=row_num, column=col_idx).value

        rows = []
        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            question_id = get_cell(row_idx, "Question ID")
            question = get_cell(row_idx, "Question")
            domain = get_cell(row_idx, "Domain")
            answer = get_cell(row_idx, "Answer")
            score = get_cell(row_idx, "Score")
            weight = get_cell(row_idx, "Weight")
            notes = get_cell(row_idx, "Notes")
            proof = get_cell(row_idx, "Proof")

            if all(v in (None, "") for v in [question_id, question, domain, answer, score, weight, notes, proof]):
                continue

            rows.append(
                {
                    "excel_row": row_idx,
                    "domain": "" if domain is None else str(domain).strip(),
                    "question_id": "" if question_id is None else str(question_id).strip(),
                    "question": "" if question is None else str(question).strip(),
                    "answer": "" if answer is None else str(answer).strip(),
                    "score": score,
                    "weight": weight,
                    "notes": "" if notes is None else str(notes),
                    "proof": "" if proof is None else str(proof),
                }
            )

        return rows

    def validate_assessment_excel(self, actor, *, assessment, excel_bytes: bytes):
        require_role(actor, "viewer")

        rows = self._read_excel_rows(excel_bytes)
        errors = []
        valid_rows = []

        for row in rows:
            row_errors = []

            if not row["question_id"]:
                row_errors.append("Question ID lipsa")
            if not row["question"]:
                row_errors.append("Question lipsa")

            if row["answer"] not in VALID_ANSWER_VALUES:
                row_errors.append("Answer invalid (acceptat: Fail, Partial, Pass, NotApplicable)")

            score = row["score"]
            if score not in (None, ""):
                try:
                    score = float(score)
                    if score < 0 or score > 100:
                        row_errors.append("Score trebuie sa fie intre 0 si 100")
                except (TypeError, ValueError):
                    row_errors.append("Score invalid")

            weight = row["weight"]
            if weight not in (None, ""):
                try:
                    weight = float(weight)
                    if weight < 0:
                        row_errors.append("Weight trebuie sa fie >= 0")
                except (TypeError, ValueError):
                    row_errors.append("Weight invalid")

            if row_errors:
                errors.append(
                    {
                        "Excel Row": row["excel_row"],
                        "Question ID": row["question_id"],
                        "Question": row["question"],
                        "Errors": " | ".join(row_errors),
                    }
                )
            else:
                valid_rows.append(row)

        return {
            "total_rows": len(rows),
            "valid_rows": len(valid_rows),
            "invalid_rows": len(errors),
            "errors": errors,
            "rows": valid_rows,
        }

    def import_assessment_excel(
        self,
        actor,
        *,
        company,
        assessment,
        excel_bytes: bytes,
        overwrite: bool = False,
    ) -> int:
        require_role(actor, "auditor")

        validation = self.validate_assessment_excel(
            actor,
            assessment=assessment,
            excel_bytes=excel_bytes,
        )

        if validation["invalid_rows"] > 0:
            raise ValueError("Fisierul Excel contine randuri invalide. Corecteaza erorile inainte de import.")

        rows = validation["rows"]

        if overwrite:
            self.session.query(Answer).filter(
                Answer.assessment_id == assessment.id
            ).delete(synchronize_session=False)

        updated_count = 0

        for item in rows:
            score = item["score"]
            if score not in (None, ""):
                score = float(score)
            else:
                score = None

            weight = item["weight"]
            if weight not in (None, ""):
                weight = float(weight)
            else:
                weight = 1.0

            self.answer_repo.upsert(
                assessment_id=assessment.id,
                question_code=item["question_id"],
                question_text=item["question"],
                domain_code=None,
                domain_name=item["domain"],
                selected_value=item["answer"],
                score=score,
                max_score=100.0,
                weight=weight,
                comment=item["notes"],
                evidence=item["proof"],
                answered_by=getattr(actor, "id", None),
                answered_at=None,
            )
            updated_count += 1

        self.session.commit()
        return updated_count

    def export_sqlite_db(self, actor) -> bytes:
        require_role(actor, "admin")

        db_path = Path("assessment.db")
        if not db_path.exists():
            raise FileNotFoundError("assessment.db nu exista.")

        return db_path.read_bytes()

    def import_sqlite_db(self, actor, db_bytes: bytes):
        require_role(actor, "admin")

        db_path = Path("assessment.db")
        backup_path = Path("assessment.db.bak")

        if db_path.exists():
            shutil.copy2(db_path, backup_path)

        db_path.write_bytes(db_bytes)
        return str(db_path)